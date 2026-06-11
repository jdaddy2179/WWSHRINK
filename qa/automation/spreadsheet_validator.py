#!/usr/bin/env python3
"""
Environment Variables spreadsheet validator
===========================================
Validates a client tab of the Environment Variables workbook against the
playbook's Phase 1 / Phase 2.1 rules — data-level checks, no browser needed.

Why not Playwright: the spreadsheet lives in SharePoint behind Okta/M365 SSO,
and validation is about *cell values/formulas*, not UI. Reading the .xlsx with
openpyxl is deterministic and repeatable.

USAGE
  1. Download "Environment Variables.xlsx" from SharePoint (your auth).
  2. python spreadsheet_validator.py "Environment Variables.xlsx" "Client X"
       (second arg = the client tab name; omit to validate "Master_Sheet")

It locates the "Key" column and the value column automatically and maps
Key -> Value, then runs the rules below. Exit code 0 = all pass, 1 = failures.

Covers test cases: TC-P1-01/03/07/08/09 (Phase 1) and TC-P2.1 Account ID format.
"""
import sys, re
from openpyxl import load_workbook

# ---- rule helpers -----------------------------------------------------------
def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip()).lower()

def digits_only(s):
    return re.sub(r'[^0-9]', '', str(s or ''))

def expected_tier(member_count):
    n = int(digits_only(member_count) or -1)
    if n < 0: return None
    if n >= 1_000_000: return "Tier 1"
    if n >= 100_000:  return "Tier 2"
    return "Tier 3"

# ---- load -------------------------------------------------------------------
def load_keymap(path, tab):
    wb = load_workbook(path, data_only=True)   # data_only -> evaluated formula values
    ws = wb[tab] if tab in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # find header row + the Key column and a value column
    key_col = val_col = header_row = None
    for ri, row in enumerate(rows[:15]):
        for ci, cell in enumerate(row or []):
            if norm(cell) == "key":
                header_row, key_col = ri, ci
        if key_col is not None:
            # value column: prefer a header literally "value", else default col C (index 2)
            for ci, cell in enumerate(rows[header_row] or []):
                if norm(cell) == "value":
                    val_col = ci
            if val_col is None:
                val_col = 2
            break
    if key_col is None:
        # fallback: assume col A = key, col C = value, no header
        key_col, val_col, header_row = 0, 2, -1
    kv = {}
    for row in rows[header_row + 1:]:
        if not row or key_col >= len(row): continue
        k = norm(row[key_col])
        if not k: continue
        v = row[val_col] if val_col < len(row) else None
        kv[k] = v
    return ws.title, kv

# ---- checks -----------------------------------------------------------------
REQUIRED_MANUAL = [
    "tenant name", "tenant id", "single-tenant or multi-tenant",
    "client member count", "tier", "tenant country code",
    "cost center", "atg", "new role or existing role",
]

def run_checks(kv):
    res = []  # (status, check, detail)
    def add(ok, check, detail=""): res.append(("PASS" if ok else "FAIL", check, detail))
    def warn(check, detail): res.append(("WARN", check, detail))

    g = lambda k: kv.get(norm(k))

    # 1. required Manual fields present & not placeholder
    for k in REQUIRED_MANUAL:
        v = g(k)
        present = v not in (None, "") and "ref[" not in norm(v)
        add(present, f"Required field present: '{k}'", "" if present else f"missing/placeholder: {v!r}")

    # 2. no leftover REF[] anywhere
    refs = [k for k, v in kv.items() if "ref[" in norm(v)]
    add(not refs, "No unresolved REF[] placeholders", f"found in: {refs}" if refs else "")

    # 3. member count numeric
    mc = g("client member count")
    mc_ok = digits_only(mc).isdigit() and digits_only(mc) != ""
    add(mc_ok, "Member count is numeric", "" if mc_ok else f"value: {mc!r}")

    # 4. Tier matches member count (TC-P1-07, the critical one)
    if mc_ok:
        exp = expected_tier(mc); got = str(g("tier") or "").strip()
        add(norm(got) == norm(exp), "Tier matches member count",
            f"members={int(digits_only(mc)):,} -> expected {exp}, sheet has {got!r}")

    # 5. Tenant ID exactly 3 chars
    tid = str(g("tenant id") or "").strip()
    add(len(tid) == 3, "Tenant ID is 3 characters", f"value: {tid!r} (len {len(tid)})")

    # 6. Cost Center 6 digits
    cc = g("cost center")
    cc_ok = re.fullmatch(r"\d{6}", digits_only(cc) or "") is not None
    add(cc_ok, "Cost Center is 6 digits", f"value: {cc!r}")

    # 7. Country code short code (2 letters)
    ccode = str(g("tenant country code") or "").strip()
    add(re.fullmatch(r"[A-Za-z]{2}", ccode) is not None, "Country code is 2-letter short code",
        f"value: {ccode!r}")

    # 8. Single/Multi-Tenant valid
    st = norm(g("single-tenant or multi-tenant"))
    add(st in ("single-tenant", "multi-tenant"), "Tenant model is Single/Multi-Tenant",
        f"value: {g('single-tenant or multi-tenant')!r}")

    # 9. Account IDs (Phase 2.1) 12 digits, if present
    for envk in ("account id", "dev account id", "qar account id", "prod account id", "hfx account id"):
        v = g(envk)
        if v not in (None, ""):
            ok = re.fullmatch(r"\d{12}", digits_only(v)) is not None
            add(ok, f"Account ID is 12 digits: '{envk}'", f"value: {v!r}")

    return res

# ---- main -------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(2)
    path = sys.argv[1]; tab = sys.argv[2] if len(sys.argv) > 2 else "Master_Sheet"
    title, kv = load_keymap(path, tab)
    print(f"Validating tab: {title}  ({len(kv)} keys read)\n")
    res = run_checks(kv)
    width = max(len(c) for _, c, _ in res)
    fails = 0
    for status, check, detail in res:
        mark = {"PASS": "✅", "FAIL": "❌", "WARN": "⚠️ "}[status]
        print(f"  {mark} {check.ljust(width)}  {detail}")
        if status == "FAIL": fails += 1
    print(f"\n{'-'*60}\n{len(res)} checks, {fails} failed.")
    sys.exit(1 if fails else 0)

if __name__ == "__main__":
    main()
